import road_mapping.model_loader as model_loader
import os
import openpyxl
import pytest

TEST_DATA_FOLDER_PATH = 'tests/data'
TEST_DATA_EMPTY_FILE_CSV = 'testdata_empty_file.csv'
TEST_DATA_EMPTY_FILE_XLSX = 'testdata_empty_file.xlsx'
TEST_DATA_UNNORMALIZED_SIMPLE = 'testdata_unnormalized_simple.xlsx'
TEST_DATA_UNNORMALIZED_HIERARCHY_DEFINED_BY_DATA_GROUPS = 'testdata_unnormalized_hierarchy_defined_by_data_groups.xlsx'
TEST_DATA_UNNORMALIZED_HIERARCHY_DEFINED_BY_LINK_TO_PARENT = 'testdata_unnormalized_hierarchy_defined_by_link_to_parent.xlsx'

def test_parseWorksheet():
    filePath = os.path.join(TEST_DATA_FOLDER_PATH, TEST_DATA_UNNORMALIZED_SIMPLE)
    workbook = openpyxl.load_workbook(filePath)
    sheet = workbook['to_import']

    sheetSummary = model_loader.xlsx_loader.parseWorksheet(sheet)

    assert(sheetSummary is not None)
    assert(sheetSummary.hasIdColumn())
    assert(sheetSummary.hasParentIdColumn())
    assert(sheetSummary.hasNameColumn())
    assert(sheetSummary.hasSummaryColumn())
    assert(sheetSummary.hasDescriptionColumn())


def test_createCapabilityFromRow():
    filePath = os.path.join(TEST_DATA_FOLDER_PATH, TEST_DATA_UNNORMALIZED_SIMPLE)
    workbook = openpyxl.load_workbook(filePath)
    sheet = workbook['to_import']

    sheetSummary = model_loader.xlsx_loader.parseWorksheet(sheet)

    capability = model_loader.xlsx_loader.createCapabilityFromRow(sheet, sheetSummary, 2)
    
    assert(capability is not None)
    assert(capability.name)
    assert(capability.summary)
    assert(capability.description)


def test_createFeatureFromRow():
    filePath = os.path.join(TEST_DATA_FOLDER_PATH, TEST_DATA_UNNORMALIZED_SIMPLE)
    workbook = openpyxl.load_workbook(filePath)
    sheet = workbook['to_import']

    sheetSummary = model_loader.xlsx_loader.parseWorksheet(sheet)

    with pytest.raises(Exception):
        model_loader.xlsx_loader.createFeatureFromRow(sheet, sheetSummary, 2)

    feature = model_loader.xlsx_loader.createFeatureFromRow(sheet, sheetSummary, 3)

    assert(feature is not None)
    assert(feature.name)
    assert(feature.summary)
    assert(feature.description)


def test_invalid_file_path():
    # input file is a folder
    filePath = os.path.join(TEST_DATA_FOLDER_PATH, '')
    with pytest.raises(Exception):
        _, _ = model_loader.xlsx_loader.loadModel(filePath)

    # input file is a different extension
    filePath = os.path.join(TEST_DATA_FOLDER_PATH, TEST_DATA_EMPTY_FILE_CSV)
    with pytest.raises(Exception):
        _, _ = model_loader.xlsx_loader.loadModel(filePath)


def test_unnormalized_hierarchy_defined_by_data_groups():
    filePath = os.path.join(TEST_DATA_FOLDER_PATH, TEST_DATA_UNNORMALIZED_HIERARCHY_DEFINED_BY_DATA_GROUPS)
    capabilities, features = model_loader.xlsx_loader.loadModel(filePath)

    assert(len(capabilities) > 0)


def test_unnormalized_hierarchy_defined_by_link_to_parent():
    filePath = os.path.join(TEST_DATA_FOLDER_PATH, TEST_DATA_UNNORMALIZED_HIERARCHY_DEFINED_BY_LINK_TO_PARENT)
    capabilities, features = model_loader.xlsx_loader.loadModel(filePath)

    assert(len(capabilities) > 0)
    assert(len(features) > 0)
