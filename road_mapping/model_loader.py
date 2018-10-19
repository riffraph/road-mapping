import openpyxl 
import os
from . import model, log, const

class worksheet_summary():
    def __init__(self):
        self.idColumn = -1
        self.nameColumn = -1
        self.summaryColumn = -1
        self.descriptionColumn = -1
        self.parentIdColumn = -1

    def hasIdColumn(self):
        return self.idColumn != -1

    def hasParentIdColumn(self):
        return self.parentIdColumn != -1

    def hasNameColumn(self):
        return self.nameColumn != -1

    def hasSummaryColumn(self):
        return self.summaryColumn != -1

    def hasDescriptionColumn(self):
        return self.descriptionColumn != -1

class xlsx_loader:

    @staticmethod
    def loadModel(filePath, logger=log.get_logger('loadModel')):
        logger.debug('xlsx_loader.loadModel({})'.format(filePath))

        if os.path.isfile(filePath) != True:
            raise Exception('model_loader: not a file')
        
        _, ext = os.path.splitext(filePath)
        if any(ext in s for s in const.SUPPORTED_FILE_EXTENSIONS) != True:
            raise Exception('model_loader: file extension {}'.format(ext))

        capabilities = []
        features = []

        ### ASSUMPTIONS ###
        # assume the data is formatted in a worksheet named 'to_import'

        # if the worksheet is not found, an exception is thrown
        workbook = openpyxl.load_workbook(filePath)
        logger.debug('workbook loaded: {}'.format(filePath))
        sheet = workbook[const.WORKSHEET_NAME]
        sheetSummary = xlsx_loader.parseWorksheet(sheet, logger)
        
        # load model objects
        for row in range(2, sheet.max_row + 1):
            rowHasParent = False

            if sheetSummary.hasParentIdColumn() == True:
                if sheet.cell(row=row, column=sheetSummary.parentIdColumn).value:
                    rowHasParent = True

            if rowHasParent == False:
                capability = xlsx_loader.createCapabilityFromRow(sheet, sheetSummary, row)

                capabilities.append(capability)
                logger.debug('Capability added: {}'.format(capability))

            else:
                feature = xlsx_loader.createFeatureFromRow(sheet, sheetSummary, row)

                features.append(feature)
                logger.debug('Feature added: {}'.format(feature))


        return capabilities, features


    @staticmethod
    def parseWorksheet(worksheet, logger=log.get_logger('parseWorksheet')):
        logger.debug('worksheet found: max_column={} max_row={}'.format(worksheet.max_column, worksheet.max_row))

        sheetSummary = worksheet_summary()

        ### ASSUMPTIONS
        # assume table has a header row, with the appropriate column titles
        # assume the minimun data for a row is an id
        # assume the table is positioned at the top left of the worksheet (cell A1)
        # assume the data is arranged as contiguous rows beneath the header

        # find the relevant columns
        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=1, column=col)
            logger.debug('column found: value={}, data_type={}'.format(cell.value, cell.data_type))
            if cell.data_type == openpyxl.cell.cell.Cell.TYPE_STRING:
                cellValue = cell.value.lower()

                if cellValue == 'id':
                    sheetSummary.idColumn = col
                    logger.info('id column found')
                elif cellValue == 'name':
                    sheetSummary.nameColumn = col
                    logger.info('name column found')
                elif cellValue == 'summary':
                    sheetSummary.summaryColumn = col
                    logger.info('summary column found')
                elif cellValue == 'description':
                    sheetSummary.descriptionColumn = col
                    logger.info('description column found')
                elif cellValue == 'parent id':
                    sheetSummary.parentIdColumn = col
                    logger.info('parent id column found')
        
        if not sheetSummary.hasIdColumn():
            raise Exception('Id column not found')

        return sheetSummary


    @staticmethod
    def createCapabilityFromRow(worksheet, worksheet_summary, row):
        capability = None

        cell = worksheet.cell(row=row, column=worksheet_summary.idColumn)
        capability = model.Capability(cell.value)

        if worksheet_summary.hasNameColumn():
            cell = worksheet.cell(row=row, column=worksheet_summary.nameColumn)
            capability.name = cell.value
        
        if worksheet_summary.hasSummaryColumn():
            cell = worksheet.cell(row=row, column=worksheet_summary.summaryColumn)
            capability.summary = cell.value

        if worksheet_summary.hasDescriptionColumn():
            cell = worksheet.cell(row=row, column=worksheet_summary.descriptionColumn)
            capability.description = cell.value

        return capability


    @staticmethod
    def createFeatureFromRow(worksheet, worksheet_summary, row):
        feature = None

        featureId = worksheet.cell(row=row, column=worksheet_summary.idColumn).value
        parentId = worksheet.cell(row=row, column=worksheet_summary.parentIdColumn).value

        if not parentId:
            raise Exception('createFeatureFromRow: feature must have a parent - {}'.format(parentId))

        feature = model.Feature(featureId, parentId)

        if worksheet_summary.hasNameColumn():
            cell = worksheet.cell(row=row, column=worksheet_summary.nameColumn)
            feature.name = cell.value
        
        if worksheet_summary.hasSummaryColumn():
            cell = worksheet.cell(row=row, column=worksheet_summary.summaryColumn)
            feature.summary = cell.value

        if worksheet_summary.hasDescriptionColumn():
            cell = worksheet.cell(row=row, column=worksheet_summary.descriptionColumn)
            feature.description = cell.value

        return feature