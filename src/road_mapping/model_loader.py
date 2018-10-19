import openpyxl 
import os
import road_mapping.model as model
import road_mapping.log as log

SUPPORTED_FILE_EXTENSIONS = ['.xlsx', '.xlsm']
WORKSHEET_NAME = 'to_import'

class xlsx_loader:
    def loadModel(filePath):
        logger = log.get_logger('model_loader')
        logger.debug('xlsx_loader.loadModel({})'.format(filePath))

        if os.path.isfile(filePath) != True:
            raise Exception('model_loader: not a file')
        
        _, ext = os.path.splitext(filePath)
        if any(ext in s for s in SUPPORTED_FILE_EXTENSIONS) != True:
            raise Exception('model_loader: file extension {}'.format(ext))

        capabilities = []
        features = []

        ### ASSUMPTIONS ###
        # assume the data is formatted in a worksheet named 'to_import'
        # assume table has a header row, with the appropriate column titles
        # assume the minimun data for a row is an id
        # assume the table is positioned at the top left of the worksheet (cell A1)
        # assume the data is arranged as contiguous rows beneath the header

        # if the worksheet is not found, an exception is thrown
        workbook = openpyxl.load_workbook(filePath)
        logger.debug('workbook loaded: {}'.format(filePath))
        sheet = workbook[WORKSHEET_NAME]
        
        idColumn = -1
        nameColumn = -1
        summaryColumn = -1
        descriptionColumn = -1
        parentIdColumn = -1

        logger.debug('sheet: max_column={}'.format(sheet.max_column))

        # find the relevant columns
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=1, column=col)
            logger.debug('column found: value={}, data_type={}'.format(cell.value, cell.data_type))
            if cell.data_type == openpyxl.cell.cell.Cell.TYPE_STRING:
                cellValue = cell.value.lower()

                if cellValue == 'id':
                    idColumn = col
                    logger.debug('id column found')
                elif cellValue == 'name':
                    nameColumn = col
                    logger.debug('name column found')
                elif cellValue == 'summary':
                    summaryColumn = col
                    logger.debug('summary column found')
                elif cellValue == 'description':
                    descriptionColumn = col
                    logger.debug('description column found')
                elif cellValue == 'parent id':
                    parentIdColumn = col
                    logger.debug('parent id column found')

            # should probably add another exit condition here

        if idColumn == -1:
            raise Exception('model_loader: Id column not found')

        # load model objects
        for row in range(2, sheet.max_row):
            hasParent = False
            parentId = -1

            if parentIdColumn != -1:
                logger.debug('parentId={}'.format(sheet.cell(row=row, column=parentIdColumn).value))
                parentId = sheet.cell(row=row, column=parentIdColumn).value
                if parentId:
                    hasParent = True

            if hasParent == False:
                cell = sheet.cell(row=row, column=idColumn)
                capability = model.Capability(cell.value)

                if nameColumn != -1:
                    cell = sheet.cell(row=row, column=nameColumn)
                    capability.name = cell.value
                
                if summaryColumn != -1:
                    cell = sheet.cell(row=row, column=summaryColumn)
                    capability.summary = cell.value

                if descriptionColumn != -1:
                    cell = sheet.cell(row=row, column=descriptionColumn)
                    capability.description = cell.value

                capabilities.append(capability)
                logger.debug('Capability added: {}'.format(capability))

            else:
                cell = sheet.cell(row=row, column=idColumn)
                feature = model.Feature(cell.value, parentId)

                if nameColumn != -1:
                    cell = sheet.cell(row=row, column=nameColumn)
                    feature.name = cell.value
                
                if summaryColumn != -1:
                    cell = sheet.cell(row=row, column=summaryColumn)
                    feature.summary = cell.value

                if descriptionColumn != -1:
                    cell = sheet.cell(row=row, column=descriptionColumn)
                    feature.description = cell.value

                features.append(feature)
                logger.debug('Feature added: {}'.format(feature))


        return capabilities, features